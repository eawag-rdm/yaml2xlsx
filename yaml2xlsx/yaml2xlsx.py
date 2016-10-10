# _*_ coding: utf-8 _*_

import os
import sys
import argparse
import yaml as yml
from openpyxl import Workbook
from functools import reduce


class YAML2XLSX(object):
    def __init__(self, arguments):
        self.parser = self.mk_parser()
        self.parameters = self.parse(arguments)
        self.data = self.read_file()

    def mk_parser(self):
        pa = argparse.ArgumentParser('yaml2xlsx', description='''
Converts a file containing YAML-documents into an xlsx-worksheet. Each
YAML-document represents one line.
''')
        pa.add_argument('infile', metavar='INPUTFILE', help=
                        'The YAML-file to convert')
        pa.add_argument('outfile', metavar='OUTPUTFILE', help=
                        'The xlsx-file to write')
        return pa
    
    def parse(self, arguments=None):
        args = vars(self.parser.parse_args(arguments))
        return args

    def read_file(self):
        in_file = self.parameters['infile']
        with open(in_file, 'rb') as f:
            data = list(yml.load_all(f))
        data = [d for d in data if d]
        return data
    
    def yaml2xlsx(self):
        out_file = self.parameters['outfile']
        in_file = self.parameters['infile']
        coldict = {}
        wb = Workbook()
        ws = wb.active
        ws.title = os.path.splitext(os.path.basename(in_file))[0]
        columns = self.collect_columns()
        for j, title in enumerate(columns):
            coldict[title] = j+1
            c = ws.cell(row=1, column=j+1, value=title)
        for row, d in enumerate(self.data):
            for colnam, val in d.iteritems():
                ws.cell(row=row+2, column=coldict[colnam], value=str(val))
        wb.save(out_file)

    def collect_columns(self):
        return list(
            set(reduce(lambda x, y: x+y, [r.keys() for r in self.data]))
        )

def main():
    C = YAML2XLSX(sys.argv[1:])
    C.yaml2xlsx()
